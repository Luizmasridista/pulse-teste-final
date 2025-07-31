#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Teste Unitário para BackupManager

Teste temporário para validar a funcionalidade do sistema de backup
conforme especificado na Sprint 3.

Este arquivo será excluído após os testes funcionarem.
"""

import pytest
import tempfile
import shutil
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook

# Importar o módulo a ser testado
import sys
sys.path.append('src')
from src.sync.backup_manager import BackupManager
from src.core.config import Config
from src.core.exceptions import BackupError


class TestBackupManager:
    """
    Testes unitários para o BackupManager.
    """
    
    def setup_method(self):
        """
        Configuração para cada teste.
        """
        # Criar diretório temporário para testes
        self.temp_dir = Path(tempfile.mkdtemp())
        self.master_folder = self.temp_dir / 'MESTRE'
        self.backup_folder = self.temp_dir / 'BACKUP'
        
        # Criar pastas
        self.master_folder.mkdir(parents=True)
        self.backup_folder.mkdir(parents=True)
        
        # Configurar caminhos para teste
        # Sobrescrever os caminhos da configuração para usar diretórios temporários
        Config.MESTRE_DIR = self.master_folder
        Config.BACKUP_DIR = self.backup_folder
        
        self.backup_manager = BackupManager()
    
    def teardown_method(self):
        """
        Limpeza após cada teste.
        """
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)
    
    def create_test_excel_file(self, file_path: Path):
        """
        Cria um arquivo Excel de teste.
        """
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Teste'
        ws['B1'] = 'Dados'
        ws['A2'] = 'Linha 1'
        ws['B2'] = 123
        wb.save(file_path)
        wb.close()
    
    def test_backup_manager_initialization(self):
        """
        Testa a inicialização do BackupManager.
        """
        assert self.backup_manager.master_folder == self.master_folder
        assert self.backup_manager.backup_folder == self.backup_folder
        assert self.backup_manager.retention_days == 30
        assert self.backup_manager.master_filename == 'planilha_consolidada.xlsx'
    
    def test_create_backup_no_master_file(self):
        """
        Testa criação de backup quando não há planilha mestre.
        """
        result = self.backup_manager.create_backup()
        assert result is None
    
    def test_create_backup_with_master_file(self):
        """
        Testa criação de backup com planilha mestre existente.
        """
        # Criar planilha mestre de teste
        master_file = self.master_folder / 'planilha_consolidada.xlsx'
        self.create_test_excel_file(master_file)
        
        # Criar backup
        backup_path = self.backup_manager.create_backup()
        
        # Verificar se backup foi criado
        assert backup_path is not None
        assert backup_path.exists()
        assert backup_path.parent == self.backup_folder
        assert 'planilha_consolidada.xlsx' in backup_path.name
    
    def test_backup_filename_format(self):
        """
        Testa o formato do nome do arquivo de backup.
        """
        filename = self.backup_manager._generate_backup_filename()
        
        # Verificar formato: YYYY-MM-DD_HH-MM-SS_planilha_consolidada.xlsx
        parts = filename.split('_')
        assert len(parts) == 4
        assert parts[2] == 'planilha'
        assert parts[3] == 'consolidada.xlsx'
        
        # Verificar se a data é válida
        date_part = f"{parts[0]}_{parts[1]}"
        try:
            datetime.strptime(date_part, "%Y-%m-%d_%H-%M-%S")
        except ValueError:
            pytest.fail("Formato de data inválido no nome do backup")
    
    def test_backup_integrity_validation(self):
        """
        Testa a validação de integridade do backup.
        """
        # Criar arquivo original
        original_file = self.master_folder / 'test.xlsx'
        self.create_test_excel_file(original_file)
        
        # Criar cópia (backup)
        backup_file = self.backup_folder / 'backup_test.xlsx'
        shutil.copy2(original_file, backup_file)
        
        # Validar integridade
        is_valid = self.backup_manager._validate_backup_integrity(original_file, backup_file)
        assert is_valid is True
    
    def test_backup_integrity_validation_corrupted(self):
        """
        Testa validação com arquivo corrompido.
        """
        # Criar arquivo original
        original_file = self.master_folder / 'test.xlsx'
        self.create_test_excel_file(original_file)
        
        # Criar arquivo corrompido
        backup_file = self.backup_folder / 'corrupted.xlsx'
        with open(backup_file, 'w') as f:
            f.write('arquivo corrompido')
        
        # Validar integridade (deve falhar)
        is_valid = self.backup_manager._validate_backup_integrity(original_file, backup_file)
        assert is_valid is False
    
    def test_list_backups(self):
        """
        Testa listagem de backups.
        """
        # Criar alguns backups de teste
        backup1 = self.backup_folder / '2025-01-29_10-30-00_planilha_consolidada.xlsx'
        backup2 = self.backup_folder / '2025-01-30_15-45-30_planilha_consolidada.xlsx'
        
        self.create_test_excel_file(backup1)
        self.create_test_excel_file(backup2)
        
        # Listar backups
        backups = self.backup_manager.list_backups()
        
        assert len(backups) == 2
        # Verificar se estão ordenados por data (mais recente primeiro)
        assert backups[0][1] > backups[1][1]
    
    def test_cleanup_old_backups(self):
        """
        Testa limpeza de backups antigos.
        """
        # Configurar retenção para 1 dia
        self.backup_manager.retention_days = 1
        
        # Criar backup antigo (2 dias atrás)
        old_date = datetime.now() - timedelta(days=2)
        old_backup_name = f"{old_date.strftime('%Y-%m-%d_%H-%M-%S')}_planilha_consolidada.xlsx"
        old_backup = self.backup_folder / old_backup_name
        self.create_test_excel_file(old_backup)
        
        # Criar backup recente
        recent_date = datetime.now()
        recent_backup_name = f"{recent_date.strftime('%Y-%m-%d_%H-%M-%S')}_planilha_consolidada.xlsx"
        recent_backup = self.backup_folder / recent_backup_name
        self.create_test_excel_file(recent_backup)
        
        # Executar limpeza
        self.backup_manager._cleanup_old_backups()
        
        # Verificar se backup antigo foi removido e recente mantido
        assert not old_backup.exists()
        assert recent_backup.exists()
    
    def test_backup_statistics(self):
        """
        Testa cálculo de estatísticas de backup.
        """
        # Criar alguns backups
        backup1 = self.backup_folder / '2025-01-29_10-30-00_planilha_consolidada.xlsx'
        backup2 = self.backup_folder / '2025-01-30_15-45-30_planilha_consolidada.xlsx'
        
        self.create_test_excel_file(backup1)
        self.create_test_excel_file(backup2)
        
        # Obter estatísticas
        stats = self.backup_manager.get_backup_statistics()
        
        assert stats['total_backups'] == 2
        assert 'oldest_backup' in stats
        assert 'newest_backup' in stats
        assert 'total_size_mb' in stats
        assert stats['total_size_mb'] > 0
    
    def test_disk_space_check(self):
        """
        Testa verificação de espaço em disco.
        """
        # Verificar com requisito baixo (deve passar)
        has_space = self.backup_manager.check_disk_space(1)  # 1 MB
        assert has_space is True
        
        # Verificar com requisito muito alto (pode falhar dependendo do sistema)
        # has_space = self.backup_manager.check_disk_space(999999999)  # 999 GB
        # assert has_space is False


def run_tests():
    """
    Executa os testes do BackupManager.
    """
    print("🧪 Iniciando testes do BackupManager...")
    
    # Executar testes
    pytest.main([__file__, '-v'])
    
    print("✅ Testes do BackupManager concluídos!")


if __name__ == '__main__':
    run_tests()