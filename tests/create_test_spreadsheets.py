"""Script para criar planilhas de teste para validação das funcionalidades.

Cria diferentes tipos de planilhas para testar:
- Planilhas válidas com dados
- Planilhas vazias
- Planilhas com diferentes formatos
- Planilhas com fórmulas e estilos
"""

import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import xlsxwriter


def create_test_directory():
    """Cria diretório de teste se não existir."""
    test_dir = Path(__file__).parent / "test_spreadsheets"
    test_dir.mkdir(exist_ok=True)
    return test_dir


def create_valid_spreadsheet_openpyxl(file_path: Path):
    """Cria planilha válida usando openpyxl.
    
    Args:
        file_path: Caminho do arquivo a ser criado.
    """
    wb = Workbook()
    
    # Primeira aba - Dados de vendas
    ws1 = wb.active
    ws1.title = "Vendas"
    
    # Cabeçalhos
    headers = ["Data", "Produto", "Quantidade", "Preço", "Total"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Dados de exemplo
    data = [
        ["2024-01-01", "Produto A", 10, 25.50, "=C2*D2"],
        ["2024-01-02", "Produto B", 5, 45.00, "=C3*D3"],
        ["2024-01-03", "Produto C", 8, 30.75, "=C4*D4"],
        ["2024-01-04", "Produto A", 15, 25.50, "=C5*D5"],
        ["2024-01-05", "Produto D", 3, 120.00, "=C6*D6"],
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=value)
    
    # Formatação de bordas
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws1.iter_rows(min_row=1, max_row=6, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
    
    # Segunda aba - Resumo
    ws2 = wb.create_sheet("Resumo")
    ws2['A1'] = "Resumo de Vendas"
    ws2['A1'].font = Font(size=16, bold=True)
    
    ws2['A3'] = "Total de Itens:"
    ws2['B3'] = "=SUM(Vendas.C2:C6)"
    
    ws2['A4'] = "Valor Total:"
    ws2['B4'] = "=SUM(Vendas.E2:E6)"
    
    # Ajustar largura das colunas
    for ws in [ws1, ws2]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)


def create_empty_spreadsheet(file_path: Path):
    """Cria planilha vazia.
    
    Args:
        file_path: Caminho do arquivo a ser criado.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha Vazia"
    # Não adiciona nenhum dado
    wb.save(file_path)


def create_large_spreadsheet(file_path: Path):
    """Cria planilha grande para teste de performance.
    
    Args:
        file_path: Caminho do arquivo a ser criado.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados Grandes"
    
    # Cabeçalhos
    headers = [f"Coluna_{i}" for i in range(1, 21)]  # 20 colunas
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Dados (1000 linhas)
    for row in range(2, 1002):
        for col in range(1, 21):
            ws.cell(row=row, column=col, value=f"Dado_{row}_{col}")
    
    wb.save(file_path)


def create_complex_spreadsheet_xlsxwriter(file_path: Path):
    """Cria planilha complexa usando xlsxwriter.
    
    Args:
        file_path: Caminho do arquivo a ser criado.
    """
    workbook = xlsxwriter.Workbook(str(file_path))
    
    # Formatos
    header_format = workbook.add_format({
        'bold': True,
        'font_color': 'white',
        'bg_color': '#366092',
        'border': 1,
        'align': 'center'
    })
    
    currency_format = workbook.add_format({
        'num_format': 'R$ #,##0.00',
        'border': 1
    })
    
    date_format = workbook.add_format({
        'num_format': 'dd/mm/yyyy',
        'border': 1
    })
    
    # Aba principal
    worksheet = workbook.add_worksheet('Dashboard')
    
    # Cabeçalhos
    headers = ['ID', 'Data', 'Cliente', 'Produto', 'Quantidade', 'Preço Unit.', 'Total']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)
    
    # Dados com fórmulas
    data = [
        [1, '2024-01-01', 'Cliente A', 'Produto X', 10, 25.50],
        [2, '2024-01-02', 'Cliente B', 'Produto Y', 5, 45.00],
        [3, '2024-01-03', 'Cliente C', 'Produto Z', 8, 30.75],
        [4, '2024-01-04', 'Cliente A', 'Produto X', 15, 25.50],
        [5, '2024-01-05', 'Cliente D', 'Produto W', 3, 120.00],
    ]
    
    for row_idx, row_data in enumerate(data, 1):
        worksheet.write(row_idx, 0, row_data[0])  # ID
        worksheet.write(row_idx, 1, row_data[1], date_format)  # Data
        worksheet.write(row_idx, 2, row_data[2])  # Cliente
        worksheet.write(row_idx, 3, row_data[3])  # Produto
        worksheet.write(row_idx, 4, row_data[4])  # Quantidade
        worksheet.write(row_idx, 5, row_data[5], currency_format)  # Preço
        # Fórmula para total
        worksheet.write_formula(row_idx, 6, f'=E{row_idx+1}*F{row_idx+1}', currency_format)
    
    # Totais
    total_row = len(data) + 2
    worksheet.write(total_row, 4, 'TOTAL:', header_format)
    worksheet.write_formula(total_row, 5, f'=SUM(F2:F{len(data)+1})', currency_format)
    worksheet.write_formula(total_row, 6, f'=SUM(G2:G{len(data)+1})', currency_format)
    
    # Gráfico
    chart = workbook.add_chart({'type': 'column'})
    chart.add_series({
        'name': 'Vendas por Produto',
        'categories': f'=Dashboard!D2:D{len(data)+1}',
        'values': f'=Dashboard!G2:G{len(data)+1}',
    })
    chart.set_title({'name': 'Vendas por Produto'})
    chart.set_x_axis({'name': 'Produtos'})
    chart.set_y_axis({'name': 'Valor (R$)'})
    
    worksheet.insert_chart('I2', chart)
    
    # Segunda aba - Análise
    analysis_sheet = workbook.add_worksheet('Análise')
    
    analysis_sheet.write('A1', 'Análise de Vendas', header_format)
    analysis_sheet.write('A3', 'Métricas:')
    analysis_sheet.write('A4', 'Total de Vendas:')
    analysis_sheet.write_formula('B4', f"=SUM(Dashboard!G2:G{len(data)+1})", currency_format)
    
    analysis_sheet.write('A5', 'Média por Venda:')
    analysis_sheet.write_formula('B5', f"=AVERAGE(Dashboard!G2:G{len(data)+1})", currency_format)
    
    analysis_sheet.write('A6', 'Maior Venda:')
    analysis_sheet.write_formula('B6', f"=MAX(Dashboard!G2:G{len(data)+1})", currency_format)
    
    # Ajustar largura das colunas
    for sheet in [worksheet, analysis_sheet]:
        sheet.set_column('A:A', 5)
        sheet.set_column('B:B', 12)
        sheet.set_column('C:C', 15)
        sheet.set_column('D:D', 15)
        sheet.set_column('E:E', 12)
        sheet.set_column('F:G', 15)
    
    workbook.close()


def create_minimal_valid_spreadsheet(file_path: Path):
    """Cria planilha mínima válida.
    
    Args:
        file_path: Caminho do arquivo a ser criado.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    
    # Apenas alguns dados básicos
    ws['A1'] = "Nome"
    ws['B1'] = "Valor"
    ws['A2'] = "Item 1"
    ws['B2'] = 100
    ws['A3'] = "Item 2"
    ws['B3'] = 200
    
    wb.save(file_path)


def create_spreadsheet_with_errors(file_path: Path):
    """Cria planilha com possíveis problemas para teste.
    
    Args:
        file_path: Caminho do arquivo a ser criado.
    """
    wb = Workbook()
    
    # Aba com dados válidos
    ws1 = wb.active
    ws1.title = "Dados Válidos"
    ws1['A1'] = "Produto"
    ws1['B1'] = "Preço"
    ws1['A2'] = "Item A"
    ws1['B2'] = 10.50
    
    # Aba completamente vazia
    ws2 = wb.create_sheet("Aba Vazia")
    # Não adiciona nenhum dado
    
    # Aba com apenas cabeçalho
    ws3 = wb.create_sheet("Só Cabeçalho")
    ws3['A1'] = "Cabeçalho"
    ws3['B1'] = "Outro Cabeçalho"
    
    wb.save(file_path)


def main():
    """Função principal para criar todas as planilhas de teste."""
    print("Criando planilhas de teste...")
    
    test_dir = create_test_directory()
    
    # Lista de planilhas para criar
    spreadsheets = [
        ("planilha_valida.xlsx", create_valid_spreadsheet_openpyxl),
        ("planilha_vazia.xlsx", create_empty_spreadsheet),
        ("planilha_grande.xlsx", create_large_spreadsheet),
        ("planilha_complexa.xlsx", create_complex_spreadsheet_xlsxwriter),
        ("planilha_minima.xlsx", create_minimal_valid_spreadsheet),
        ("planilha_com_problemas.xlsx", create_spreadsheet_with_errors),
    ]
    
    created_files = []
    
    for filename, create_func in spreadsheets:
        file_path = test_dir / filename
        try:
            create_func(file_path)
            created_files.append(file_path)
            print(f"✓ Criado: {filename}")
        except Exception as e:
            print(f"✗ Erro ao criar {filename}: {e}")
    
    # Criar alguns arquivos não-Excel para teste
    non_excel_files = [
        ("arquivo.txt", "Este é um arquivo de texto."),
        ("dados.csv", "Nome,Idade,Cidade\nJoão,30,São Paulo\nMaria,25,Rio de Janeiro"),
        ("config.json", '{"configuracao": "valor"}'),
    ]
    
    for filename, content in non_excel_files:
        file_path = test_dir / filename
        try:
            file_path.write_text(content, encoding='utf-8')
            created_files.append(file_path)
            print(f"✓ Criado: {filename}")
        except Exception as e:
            print(f"✗ Erro ao criar {filename}: {e}")
    
    # Criar subdiretório com planilhas
    sub_dir = test_dir / "subdirectory"
    sub_dir.mkdir(exist_ok=True)
    
    sub_file = sub_dir / "planilha_subdir.xlsx"
    try:
        create_minimal_valid_spreadsheet(sub_file)
        created_files.append(sub_file)
        print(f"✓ Criado: subdirectory/planilha_subdir.xlsx")
    except Exception as e:
        print(f"✗ Erro ao criar planilha em subdiretório: {e}")
    
    print(f"\nTotal de arquivos criados: {len(created_files)}")
    print(f"Diretório de teste: {test_dir}")
    
    # Listar arquivos criados
    print("\nArquivos criados:")
    for file_path in created_files:
        size = file_path.stat().st_size if file_path.exists() else 0
        print(f"  {file_path.name} ({size:,} bytes)")


if __name__ == "__main__":
    main()