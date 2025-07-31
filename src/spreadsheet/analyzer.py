"""Analisador de estrutura de planilhas.

Este módulo implementa a análise detalhada da estrutura das planilhas,
incluindo extração de cabeçalhos, mapeamento de estilos e detecção de fórmulas.
"""

import os
import logging
from pathlib import Path
from typing import List, Dict, Optional, Any, Set, Tuple
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum

try:
    import openpyxl
    from openpyxl.styles import Font, Fill, Border, Alignment
    import pandas as pd
except ImportError as e:
    raise ImportError(f"Dependências necessárias não encontradas: {e}")

try:
    from ..core import get_logger, config
    from ..core.exceptions import AnalysisException
except ImportError:
    # Fallback para quando executado diretamente ou em testes
    def get_logger(name: str) -> logging.Logger:
        logger = logging.getLogger(name)
        if not logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
            handler.setFormatter(formatter)
            logger.addHandler(handler)
            logger.setLevel(logging.INFO)
        return logger
    
    class Config:
        def __init__(self):
            self.max_analysis_cells = 10000
            self.analysis_timeout = 300  # 5 minutos
    
    config = Config()
    
    class AnalysisException(Exception):
        """Exceção para erros de análise."""
        pass

try:
    from .scanner import SpreadsheetInfo
    from .validator import SpreadsheetValidationResult
except ImportError:
    # Fallback para quando executado diretamente
    @dataclass
    class SpreadsheetInfo:
        name: str
        path: Path
        size: int
        modified_date: datetime
        sheet_count: int = 0
        
    @dataclass
    class SpreadsheetValidationResult:
        is_valid: bool
        errors: List[str] = field(default_factory=list)


class CellType(Enum):
    """Tipos de célula identificados."""
    HEADER = "header"
    DATA = "data"
    FORMULA = "formula"
    EMPTY = "empty"
    MERGED = "merged"


@dataclass
class CellStyle:
    """Informações de estilo de uma célula."""
    font_name: Optional[str] = None
    font_size: Optional[int] = None
    font_bold: bool = False
    font_italic: bool = False
    font_color: Optional[str] = None
    fill_color: Optional[str] = None
    border_style: Optional[str] = None
    alignment_horizontal: Optional[str] = None
    alignment_vertical: Optional[str] = None
    number_format: Optional[str] = None


@dataclass
class CellInfo:
    """Informações detalhadas de uma célula."""
    row: int
    column: int
    address: str
    value: Any
    cell_type: CellType
    style: CellStyle
    formula: Optional[str] = None
    is_merged: bool = False
    merge_range: Optional[str] = None


@dataclass
class SheetAnalysis:
    """Análise de uma aba da planilha."""
    name: str
    row_count: int
    column_count: int
    headers: List[str] = field(default_factory=list)
    header_row: Optional[int] = None
    data_range: Optional[str] = None
    formulas: List[CellInfo] = field(default_factory=list)
    merged_cells: List[str] = field(default_factory=list)
    styles_map: Dict[str, CellStyle] = field(default_factory=dict)
    visual_elements: Dict[str, Any] = field(default_factory=dict)
    data_types: Dict[str, str] = field(default_factory=dict)


@dataclass
class SpreadsheetAnalysis:
    """Análise completa de uma planilha."""
    spreadsheet_info: SpreadsheetInfo
    sheets: List[SheetAnalysis] = field(default_factory=list)
    global_styles: Dict[str, CellStyle] = field(default_factory=dict)
    has_formulas: bool = False
    has_merged_cells: bool = False
    complexity_score: int = 0
    analysis_timestamp: Optional[str] = None


class SpreadsheetAnalyzer:
    """Analisador de estrutura de planilhas.
    
    Implementa o Passo 4 dos requisitos técnicos:
    - 4.1: Carregar planilhas
    - 4.2: Extrair cabeçalhos
    - 4.3: Mapear estilos de células
    - 4.4: Detectar fórmulas
    - 4.5: Catalogar elementos visuais
    """
    
    def __init__(self):
        """Inicializa o analisador."""
        self.logger = get_logger(__name__)
        
    def analyze_spreadsheet(self, spreadsheet_info: SpreadsheetInfo) -> SpreadsheetAnalysis:
        """Analisa uma planilha completa.
        
        Args:
            spreadsheet_info: Informações da planilha a ser analisada.
            
        Returns:
            SpreadsheetAnalysis com análise completa.
        """
        self.logger.info(f"Analisando estrutura da planilha: {spreadsheet_info.name}")
        
        try:
            # Carregar workbook com openpyxl para análise detalhada
            workbook = openpyxl.load_workbook(spreadsheet_info.path, data_only=False)
            
            analysis = SpreadsheetAnalysis(
                spreadsheet_info=spreadsheet_info,
                analysis_timestamp=pd.Timestamp.now().isoformat()
            )
            
            # Analisar cada aba
            for sheet_name in workbook.sheetnames:
                sheet_analysis = self._analyze_sheet(workbook[sheet_name], sheet_name)
                analysis.sheets.append(sheet_analysis)
                
                # Atualizar flags globais
                if sheet_analysis.formulas:
                    analysis.has_formulas = True
                if sheet_analysis.merged_cells:
                    analysis.has_merged_cells = True
                    
            # Calcular score de complexidade
            analysis.complexity_score = self._calculate_complexity_score(analysis)
            
            # Extrair estilos globais
            analysis.global_styles = self._extract_global_styles(analysis)
            
            self.logger.info(
                f"Análise concluída: {len(analysis.sheets)} abas, "
                f"complexidade: {analysis.complexity_score}"
            )
            
            return analysis
            
        except Exception as e:
            self.logger.error(f"Erro ao analisar planilha {spreadsheet_info.name}: {e}")
            raise AnalysisException(f"Falha na análise: {str(e)}")
            
    def _analyze_sheet(self, worksheet, sheet_name: str) -> SheetAnalysis:
        """Analisa uma aba específica.
        
        Args:
            worksheet: Objeto worksheet do openpyxl.
            sheet_name: Nome da aba.
            
        Returns:
            SheetAnalysis com análise da aba.
        """
        self.logger.debug(f"Analisando aba: {sheet_name}")
        
        analysis = SheetAnalysis(
            name=sheet_name,
            row_count=worksheet.max_row,
            column_count=worksheet.max_column
        )
        
        # Detectar cabeçalhos
        analysis.headers, analysis.header_row = self._detect_headers(worksheet)
        
        # Definir range de dados
        if analysis.header_row:
            start_row = analysis.header_row + 1
            analysis.data_range = f"A{start_row}:{openpyxl.utils.get_column_letter(analysis.column_count)}{analysis.row_count}"
        
        # Analisar células
        analysis.formulas = self._extract_formulas(worksheet)
        analysis.merged_cells = [str(range_) for range_ in worksheet.merged_cells.ranges]
        analysis.styles_map = self._map_cell_styles(worksheet)
        
        # Catalogar elementos visuais
        analysis.visual_elements = self._catalog_visual_elements(worksheet)
        
        # Detectar tipos de dados
        analysis.data_types = self._detect_data_types(worksheet, analysis.header_row)
        
        return analysis
        
    def _detect_headers(self, worksheet) -> Tuple[List[str], Optional[int]]:
        """Detecta cabeçalhos na planilha.
        
        Args:
            worksheet: Objeto worksheet do openpyxl.
            
        Returns:
            Tupla com (lista de cabeçalhos, linha do cabeçalho).
        """
        headers = []
        header_row = None
        
        # Verificar as primeiras 5 linhas em busca de cabeçalhos
        for row_num in range(1, min(6, worksheet.max_row + 1)):
            row_values = []
            has_text = False
            
            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                value = cell.value
                
                if value is not None:
                    if isinstance(value, str) and value.strip():
                        has_text = True
                        row_values.append(value.strip())
                    else:
                        row_values.append(str(value) if value is not None else "")
                else:
                    row_values.append("")
                    
            # Se a linha tem texto e não está vazia, pode ser cabeçalho
            if has_text and any(val for val in row_values):
                # Verificar se parece com cabeçalho (texto, sem números puros)
                text_count = sum(1 for val in row_values if val and not str(val).replace('.', '').replace(',', '').isdigit())
                
                if text_count >= len([v for v in row_values if v]) * 0.7:  # 70% texto
                    headers = [val for val in row_values if val]
                    header_row = row_num
                    break
                    
        return headers, header_row
        
    def _extract_formulas(self, worksheet) -> List[CellInfo]:
        """Extrai fórmulas da planilha.
        
        Args:
            worksheet: Objeto worksheet do openpyxl.
            
        Returns:
            Lista de CellInfo com fórmulas encontradas.
        """
        formulas = []
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Formula
                    cell_info = CellInfo(
                        row=cell.row,
                        column=cell.column,
                        address=cell.coordinate,
                        value=cell.value,
                        cell_type=CellType.FORMULA,
                        style=self._extract_cell_style(cell),
                        formula=cell.value
                    )
                    formulas.append(cell_info)
                    
        return formulas
        
    def _map_cell_styles(self, worksheet) -> Dict[str, CellStyle]:
        """Mapeia estilos de células.
        
        Args:
            worksheet: Objeto worksheet do openpyxl.
            
        Returns:
            Dicionário mapeando coordenadas para estilos.
        """
        styles_map = {}
        
        # Analisar apenas células com conteúdo ou estilo especial
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None or self._has_custom_style(cell):
                    style = self._extract_cell_style(cell)
                    styles_map[cell.coordinate] = style
                    
        return styles_map
        
    def _extract_cell_style(self, cell) -> CellStyle:
        """Extrai estilo de uma célula.
        
        Args:
            cell: Célula do openpyxl.
            
        Returns:
            CellStyle com informações de estilo.
        """
        style = CellStyle()
        
        # Font
        if cell.font:
            style.font_name = cell.font.name
            style.font_size = cell.font.size
            style.font_bold = cell.font.bold or False
            style.font_italic = cell.font.italic or False
            if cell.font.color and hasattr(cell.font.color, 'rgb'):
                style.font_color = str(cell.font.color.rgb)
                
        # Fill
        if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
            if hasattr(cell.fill.start_color, 'rgb'):
                style.fill_color = str(cell.fill.start_color.rgb)
                
        # Border
        if cell.border and any([cell.border.left, cell.border.right, cell.border.top, cell.border.bottom]):
            style.border_style = "custom"
            
        # Alignment
        if cell.alignment:
            style.alignment_horizontal = cell.alignment.horizontal
            style.alignment_vertical = cell.alignment.vertical
            
        # Number format
        style.number_format = cell.number_format
        
        return style
        
    def _has_custom_style(self, cell) -> bool:
        """Verifica se a célula tem estilo customizado.
        
        Args:
            cell: Célula do openpyxl.
            
        Returns:
            True se tem estilo customizado.
        """
        # Verificar se tem formatação diferente do padrão
        if cell.font and (cell.font.bold or cell.font.italic or cell.font.size != 11):
            return True
        if cell.fill and cell.fill.start_color and str(cell.fill.start_color.rgb) != '00000000':
            return True
        if cell.border and any([cell.border.left, cell.border.right, cell.border.top, cell.border.bottom]):
            return True
        if cell.alignment and (cell.alignment.horizontal or cell.alignment.vertical):
            return True
            
        return False
        
    def _catalog_visual_elements(self, worksheet) -> Dict[str, Any]:
        """Cataloga elementos visuais da planilha.
        
        Args:
            worksheet: Objeto worksheet do openpyxl.
            
        Returns:
            Dicionário com elementos visuais encontrados.
        """
        elements = {
            'charts': [],
            'images': [],
            'shapes': [],
            'conditional_formatting': [],
            'data_validation': [],
            'hyperlinks': []
        }
        
        # Charts
        if hasattr(worksheet, '_charts'):
            elements['charts'] = [{'type': type(chart).__name__} for chart in worksheet._charts]
            
        # Images
        if hasattr(worksheet, '_images'):
            elements['images'] = [{'anchor': str(img.anchor)} for img in worksheet._images]
            
        # Conditional formatting
        if worksheet.conditional_formatting:
            for cf in worksheet.conditional_formatting:
                elements['conditional_formatting'].append({
                    'range': str(cf.sqref),
                    'rules_count': len(cf.cfRule)
                })
                
        # Data validation
        if worksheet.data_validations:
            for dv in worksheet.data_validations.dataValidation:
                elements['data_validation'].append({
                    'range': str(dv.sqref),
                    'type': dv.type
                })
                
        # Hyperlinks
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    elements['hyperlinks'].append({
                        'cell': cell.coordinate,
                        'target': cell.hyperlink.target
                    })
                    
        return elements
        
    def _detect_data_types(self, worksheet, header_row: Optional[int]) -> Dict[str, str]:
        """Detecta tipos de dados nas colunas.
        
        Args:
            worksheet: Objeto worksheet do openpyxl.
            header_row: Linha do cabeçalho.
            
        Returns:
            Dicionário mapeando colunas para tipos de dados.
        """
        data_types = {}
        
        if not header_row:
            return data_types
            
        # Analisar algumas linhas de dados para detectar tipos
        sample_rows = min(10, worksheet.max_row - header_row)
        
        for col_num in range(1, worksheet.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            
            # Coletar amostras de valores
            values = []
            for row_num in range(header_row + 1, header_row + sample_rows + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                if cell.value is not None:
                    values.append(cell.value)
                    
            # Determinar tipo predominante
            if values:
                data_type = self._infer_data_type(values)
                data_types[col_letter] = data_type
                
        return data_types
        
    def _infer_data_type(self, values: List[Any]) -> str:
        """Infere o tipo de dados de uma lista de valores.
        
        Args:
            values: Lista de valores para análise.
            
        Returns:
            String representando o tipo de dados.
        """
        if not values:
            return "empty"
            
        # Contar tipos
        type_counts = {
            'number': 0,
            'date': 0,
            'text': 0,
            'boolean': 0
        }
        
        for value in values:
            if isinstance(value, (int, float)):
                type_counts['number'] += 1
            elif isinstance(value, pd.Timestamp) or str(value).count('/') == 2 or str(value).count('-') == 2:
                type_counts['date'] += 1
            elif isinstance(value, bool):
                type_counts['boolean'] += 1
            else:
                type_counts['text'] += 1
                
        # Retornar tipo predominante
        return max(type_counts, key=type_counts.get)
        
    def _calculate_complexity_score(self, analysis: SpreadsheetAnalysis) -> int:
        """Calcula score de complexidade da planilha.
        
        Args:
            analysis: Análise da planilha.
            
        Returns:
            Score de complexidade (0-100).
        """
        score = 0
        
        # Pontos por número de abas
        score += min(len(analysis.sheets) * 5, 20)
        
        # Pontos por fórmulas
        total_formulas = sum(len(sheet.formulas) for sheet in analysis.sheets)
        score += min(total_formulas * 2, 30)
        
        # Pontos por células mescladas
        total_merged = sum(len(sheet.merged_cells) for sheet in analysis.sheets)
        score += min(total_merged * 3, 20)
        
        # Pontos por elementos visuais
        for sheet in analysis.sheets:
            visual_count = sum(len(elements) for elements in sheet.visual_elements.values())
            score += min(visual_count * 2, 15)
            
        # Pontos por estilos customizados
        total_styles = sum(len(sheet.styles_map) for sheet in analysis.sheets)
        score += min(total_styles // 10, 15)
        
        return min(score, 100)
        
    def _extract_global_styles(self, analysis: SpreadsheetAnalysis) -> Dict[str, CellStyle]:
        """Extrai estilos globais mais comuns.
        
        Args:
            analysis: Análise da planilha.
            
        Returns:
            Dicionário com estilos globais.
        """
        # Por simplicidade, retorna estilos da primeira aba
        if analysis.sheets:
            return analysis.sheets[0].styles_map
        return {}