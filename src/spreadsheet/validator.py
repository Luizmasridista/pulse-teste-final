"""Validador de planilhas subordinadas.

Este módulo implementa a validação de integridade e verificação
de conteúdo das planilhas encontradas pelo scanner.
"""

import os
import logging
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Any
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum

try:
    import openpyxl
    import pandas as pd
except ImportError as e:
    raise ImportError(f"Dependências necessárias não encontradas: {e}")

try:
    from ..core import get_logger, config
    from ..core.exceptions import ValidationException
except ImportError:
    # Fallback para quando executado diretamente ou em testes
    def get_logger(name: str) -> logging.Logger:
        logger = logging.getLogger(name)
        if not logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
            handler.setFormatter(formatter)
            logger.addHandler(handler)
            logger.setLevel(logging.INFO)
        return logger
    
    class Config:
        def __init__(self):
            self.max_file_size = 100 * 1024 * 1024  # 100MB
            self.min_file_size = 1024  # 1KB
            self.supported_extensions = ['.xlsx', '.xls']
    
    config = Config()
    
    class ValidationException(Exception):
        """Exceção para erros de validação."""
        pass

from .scanner import SpreadsheetInfo


class ValidationStatus(Enum):
    """Status de validação de uma planilha."""
    VALID = "valid"
    INVALID = "invalid"
    WARNING = "warning"
    ERROR = "error"


@dataclass
class SpreadsheetValidationResult:
    """Resultado da validação de uma planilha."""
    file_path: str
    status: ValidationStatus
    errors: List[str]
    warnings: List[str]
    metadata: dict = None
    
    def __post_init__(self):
        """Inicialização pós-criação."""
        if self.metadata is None:
            self.metadata = {}
    
    @property
    def has_errors(self) -> bool:
        """Verifica se há erros."""
        return len(self.errors) > 0
    
    @property
    def has_warnings(self) -> bool:
        """Verifica se há avisos."""
        return len(self.warnings) > 0
    
    @property
    def is_valid(self) -> bool:
        """Verifica se é válido."""
        return self.status == ValidationStatus.VALID
    
    def __str__(self) -> str:
        """Representação string."""
        from pathlib import Path
        filename = Path(self.file_path).name
        return f"{filename}: {self.status.value.upper()}"
    
    def __repr__(self) -> str:
        """Representação repr."""
        from pathlib import Path
        filename = Path(self.file_path).name
        return f"SpreadsheetValidationResult(file='{filename}', status={self.status})"


class SpreadsheetValidator:
    """Validador de planilhas subordinadas.
    
    Implementa o Passo 3 dos requisitos técnicos:
    - 3.1: Validar integridade das planilhas
    - 3.2: Verificar se não estão vazias
    - 3.3: Logging dos resultados
    """
    
    def __init__(self):
        """Inicializa o validador."""
        self.logger = get_logger(__name__)
        
    def validate_file(self, file_path: str) -> SpreadsheetValidationResult:
        """Valida um arquivo de planilha.
        
        Args:
            file_path: Caminho para o arquivo a ser validado.
            
        Returns:
            SpreadsheetValidationResult com o resultado da validação.
        """
        from pathlib import Path
        from datetime import datetime
        
        # Criar SpreadsheetInfo a partir do caminho
        path = Path(file_path)
        if not path.exists():
            return SpreadsheetValidationResult(
                file_path=file_path,
                status=ValidationStatus.ERROR,
                errors=["Arquivo não encontrado"],
                warnings=[]
            )
        
        stat = path.stat()
        spreadsheet_info = SpreadsheetInfo(
            name=path.name,
            path=path,
            size=stat.st_size,
            modified_date=datetime.fromtimestamp(stat.st_mtime),
            extension=path.suffix
        )
        
        return self.validate_spreadsheet(spreadsheet_info)
    
    def validate_multiple_files(self, file_paths: List[str]) -> Dict[str, SpreadsheetValidationResult]:
        """Valida múltiplos arquivos de planilha.
        
        Args:
            file_paths: Lista de caminhos para os arquivos
            
        Returns:
            Dicionário com resultados da validação para cada arquivo
        """
        if not file_paths:
            return {}
            
        results = {}
        for file_path in file_paths:
            try:
                result = self.validate_file(file_path)
                results[file_path] = result
            except Exception as e:
                # Em caso de erro, criar um resultado de erro
                results[file_path] = SpreadsheetValidationResult(
                    file_path=file_path,
                    status=ValidationStatus.ERROR,
                    errors=[f"Erro ao validar arquivo: {str(e)}"],
                    warnings=[]
                )
        return results
    
    def _is_excel_file(self, file_path: str) -> bool:
        """Verifica se o arquivo é um arquivo Excel válido.
        
        Args:
            file_path: Caminho para o arquivo
            
        Returns:
            True se for um arquivo Excel (.xlsx ou .xls)
        """
        from pathlib import Path
        extension = Path(file_path).suffix.lower()
        return extension in ['.xlsx', '.xls']
    
    def _validate_xlsx_file(self, file_path: str) -> SpreadsheetValidationResult:
        """Valida especificamente um arquivo XLSX.
        
        Args:
            file_path: Caminho para o arquivo XLSX
            
        Returns:
            Resultado da validação
        """
        import openpyxl
        errors = []
        warnings = []
        
        # Verificar tamanho do arquivo (incluindo arquivos grandes)
        size_errors, size_warnings = self._check_file_size(file_path)
        errors.extend(size_errors)
        warnings.extend(size_warnings)
        
        # Se há erros de tamanho, retornar resultado inválido
        if size_errors:
            return SpreadsheetValidationResult(
                file_path=file_path,
                status=ValidationStatus.INVALID,
                errors=errors,
                warnings=warnings
            )
        
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            
            total_rows = 0
            total_columns = 0
            sheets_count = len(workbook.sheetnames)
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                total_rows += worksheet.max_row
                total_columns += worksheet.max_column

            # Verificar se as planilhas estão vazias (apenas cabeçalho ou menos)
            if total_rows <= sheets_count:  # Se total_rows <= número de sheets, significa que cada sheet tem no máximo 1 linha
                errors.append("Planilha vazia ou contém apenas cabeçalhos")
                return SpreadsheetValidationResult(
                    file_path=file_path,
                    status=ValidationStatus.INVALID,
                    errors=errors,
                    warnings=warnings
                )

            metadata = {
                'sheets_count': sheets_count,
                'total_rows': total_rows,
                'total_columns': total_columns,
                'has_data': total_rows > 0
            }

            return SpreadsheetValidationResult(
                file_path=file_path,
                status=ValidationStatus.VALID,
                errors=errors,
                warnings=warnings,
                metadata=metadata
            )
            
        except Exception as e:
            errors.append(f"Erro ao validar arquivo XLSX: {str(e)}")
            return SpreadsheetValidationResult(
                file_path=file_path,
                status=ValidationStatus.ERROR,
                errors=errors,
                warnings=warnings
            )
    
    def _validate_xls_file(self, file_path: str) -> SpreadsheetValidationResult:
        """Valida especificamente um arquivo XLS.
        
        Args:
            file_path: Caminho para o arquivo XLS
            
        Returns:
            Resultado da validação
        """
        import xlrd
        errors = []
        warnings = []
        
        # Verificar tamanho do arquivo (incluindo arquivos grandes)
        size_errors, size_warnings = self._check_file_size(file_path)
        errors.extend(size_errors)
        warnings.extend(size_warnings)
        
        # Se há erros de tamanho, retornar resultado inválido
        if size_errors:
            return SpreadsheetValidationResult(
                file_path=file_path,
                status=ValidationStatus.INVALID,
                errors=errors,
                warnings=warnings
            )
        
        try:
            workbook = xlrd.open_workbook(file_path)
            
            total_rows = 0
            total_columns = 0
            sheets_count = workbook.nsheets
            
            for sheet_index in range(workbook.nsheets):
                worksheet = workbook.sheet_by_index(sheet_index)
                total_rows += worksheet.nrows
                total_columns += worksheet.ncols
            
            metadata = {
                'sheets_count': sheets_count,
                'total_rows': total_rows,
                'total_columns': total_columns,
                'has_data': total_rows > 0
            }
            
            return SpreadsheetValidationResult(
                file_path=file_path,
                status=ValidationStatus.VALID,
                errors=errors,
                warnings=warnings,
                metadata=metadata
            )
            
        except Exception as e:
            errors.append(f"Erro ao validar arquivo XLS: {str(e)}")
            return SpreadsheetValidationResult(
                file_path=file_path,
                status=ValidationStatus.ERROR,
                errors=errors,
                warnings=warnings
            )
    
    def validate_spreadsheet(self, spreadsheet_info: SpreadsheetInfo) -> SpreadsheetValidationResult:
        """Valida uma planilha individual.
        
        Args:
            spreadsheet_info: Informações da planilha a ser validada.
            
        Returns:
            SpreadsheetValidationResult com o resultado da validação.
        """
        self.logger.info(f"Validando planilha: {spreadsheet_info.name}")
        
        errors = []
        warnings = []
        details = {}
        
        try:
            # Verificar se o arquivo existe
            if not spreadsheet_info.path.exists():
                errors.append(f"Arquivo não encontrado: {spreadsheet_info.path}")
                return SpreadsheetValidationResult(
                    file_path=str(spreadsheet_info.path),
                    status=ValidationStatus.ERROR,
                    errors=errors,
                    warnings=warnings
                )
                
            # Verificar extensão primeiro
            if spreadsheet_info.extension.lower() not in ['.xlsx', '.xls']:
                errors.append(f"Extensão não suportada: {spreadsheet_info.extension}")
                return SpreadsheetValidationResult(
                    file_path=str(spreadsheet_info.path),
                    status=ValidationStatus.INVALID,
                    errors=errors,
                    warnings=warnings
                )
                
            # Verificar se o arquivo não está vazio
            if spreadsheet_info.size == 0:
                errors.append("Arquivo está vazio (0 bytes)")
                return SpreadsheetValidationResult(
                    file_path=str(spreadsheet_info.path),
                    status=ValidationStatus.INVALID,
                    errors=errors,
                    warnings=warnings
                )
                
            # Verificar tamanho mínimo
            if spreadsheet_info.size < 1024:  # Menor que 1KB
                errors.append("Arquivo muito pequeno para ser uma planilha válida")
                return SpreadsheetValidationResult(
                    file_path=str(spreadsheet_info.path),
                    status=ValidationStatus.INVALID,
                    errors=errors,
                    warnings=warnings
                )
                
            # Verificar tamanho do arquivo (incluindo arquivos grandes)
            size_errors, size_warnings = self._check_file_size(str(spreadsheet_info.path))
            errors.extend(size_errors)
            warnings.extend(size_warnings)
            
            # Se há erros de tamanho, retornar resultado inválido
            if size_errors:
                return SpreadsheetValidationResult(
                    file_path=str(spreadsheet_info.path),
                    status=ValidationStatus.INVALID,
                    errors=errors,
                    warnings=warnings
                )
                
            # Se for arquivo Excel, usar métodos específicos
            if self._is_excel_file(str(spreadsheet_info.path)):
                if spreadsheet_info.extension.lower() == '.xlsx':
                    return self._validate_xlsx_file(str(spreadsheet_info.path))
                elif spreadsheet_info.extension.lower() == '.xls':
                    return self._validate_xls_file(str(spreadsheet_info.path))
                
            # Tentar carregar a planilha
            try:
                # Carregar todas as abas da planilha
                excel_file = pd.ExcelFile(spreadsheet_info.path)
                sheet_names = excel_file.sheet_names
                sheet_count = len(sheet_names)
                
                details['sheet_names'] = sheet_names
                details['sheet_count'] = sheet_count
                
                if sheet_count == 0:
                    errors.append("Planilha não contém abas")
                    return SpreadsheetValidationResult(
                        file_path=str(spreadsheet_info.path),
                        status=ValidationStatus.ERROR,
                        errors=errors,
                        warnings=warnings,
                        metadata=details
                    )
                    
                # Analisar cada aba
                total_rows = 0
                total_columns = 0
                has_data = False
                empty_sheets = []
                
                for sheet_name in sheet_names:
                    try:
                        df = pd.read_excel(spreadsheet_info.path, sheet_name=sheet_name)
                        
                        rows = len(df)
                        cols = len(df.columns)
                        
                        total_rows += rows
                        total_columns = max(total_columns, cols)
                        
                        # Verificar se a aba tem dados
                        if rows > 0 and cols > 0 and not df.empty:
                            # Verificar se há dados não-nulos
                            non_null_count = df.count().sum()
                            if non_null_count > 0:
                                has_data = True
                            else:
                                empty_sheets.append(sheet_name)
                        else:
                            empty_sheets.append(sheet_name)
                            
                        details[f'sheet_{sheet_name}'] = {
                            'rows': rows,
                            'columns': cols,
                            'non_null_cells': df.count().sum() if not df.empty else 0
                        }
                        
                    except Exception as e:
                        errors.append(f"Erro ao ler aba '{sheet_name}': {str(e)}")
                        
                # Verificar se há abas vazias
                if empty_sheets:
                    if len(empty_sheets) == sheet_count:
                        errors.append("Todas as abas estão vazias")
                    else:
                        warnings.append(f"Abas vazias encontradas: {', '.join(empty_sheets)}")
                        
                # Determinar status
                if errors:
                    status = ValidationStatus.ERROR
                elif not has_data:
                    status = ValidationStatus.INVALID
                    errors.append("Planilha não contém dados válidos")
                elif warnings:
                    status = ValidationStatus.WARNING
                else:
                    status = ValidationStatus.VALID
                    
                # Adicionar metadados
                details['total_rows'] = total_rows
                details['total_columns'] = total_columns
                details['has_data'] = has_data
                details['sheets_count'] = sheet_count
                
                result = SpreadsheetValidationResult(
                    file_path=str(spreadsheet_info.path),
                    status=status,
                    errors=errors,
                    warnings=warnings,
                    metadata=details
                )
                
                # Atualizar o spreadsheet_info
                spreadsheet_info.is_valid = (status == ValidationStatus.VALID)
                if errors:
                    spreadsheet_info.error_message = "; ".join(errors)
                    
                self._log_validation_result(result)
                return result
                
            except Exception as e:
                error_msg = f"Erro ao carregar planilha: {str(e)}"
                errors.append(error_msg)
                return SpreadsheetValidationResult(
                    file_path=str(spreadsheet_info.path),
                    status=ValidationStatus.ERROR,
                    errors=errors,
                    warnings=warnings,
                    metadata=details
                )
                
        except Exception as e:
            error_msg = f"Erro inesperado durante validação: {str(e)}"
            errors.append(error_msg)
            return SpreadsheetValidationResult(
                file_path=str(spreadsheet_info.path),
                status=ValidationStatus.ERROR,
                errors=errors,
                warnings=warnings
            )
            
    def validate_multiple(self, spreadsheets: List[SpreadsheetInfo]) -> List[SpreadsheetValidationResult]:
        """Valida múltiplas planilhas.
        
        Args:
            spreadsheets: Lista de planilhas a serem validadas.
            
        Returns:
            Lista de resultados de validação.
        """
        self.logger.info(f"Iniciando validação de {len(spreadsheets)} planilhas")
        
        results = []
        for spreadsheet in spreadsheets:
            try:
                result = self.validate_spreadsheet(spreadsheet)
                results.append(result)
            except Exception as e:
                self.logger.error(f"Erro ao validar {spreadsheet.name}: {e}")
                error_result = self._create_error_result(
                    spreadsheet, 
                    [f"Erro durante validação: {str(e)}"]
                )
                results.append(error_result)
                
        self._log_validation_summary(results)
        return results
        
    def get_validation_summary(self, results: List[SpreadsheetValidationResult]) -> Dict[str, any]:
        """Gera resumo das validações.
        
        Args:
            results: Lista de resultados de validação.
            
        Returns:
            Dicionário com estatísticas de validação.
        """
        total = len(results)
        valid = len([r for r in results if r.status == ValidationStatus.VALID])
        invalid = len([r for r in results if r.status == ValidationStatus.INVALID])
        errors = len([r for r in results if r.status == ValidationStatus.ERROR])
        warnings = len([r for r in results if r.status == ValidationStatus.WARNING])
        
        empty_files = len([r for r in results if r.is_empty])
        total_sheets = sum(r.sheet_count for r in results)
        total_rows = sum(r.row_count for r in results)
        
        summary = {
            'total_files': total,
            'valid_files': valid,
            'invalid_files': invalid,
            'error_files': errors,
            'warning_files': warnings,
            'empty_files': empty_files,
            'total_sheets': total_sheets,
            'total_rows': total_rows,
            'success_rate': (valid / total * 100) if total > 0 else 0
        }
        
        return summary
        
    def _create_error_result(self, spreadsheet_info: SpreadsheetInfo, 
                           errors: List[str], details: Dict = None) -> SpreadsheetValidationResult:
        """Cria um resultado de erro."""
        return SpreadsheetValidationResult(
            spreadsheet_info=spreadsheet_info,
            status=ValidationStatus.ERROR,
            is_empty=True,
            sheet_count=0,
            row_count=0,
            column_count=0,
            has_data=False,
            errors=errors,
            warnings=[],
            details=details or {}
        )
        
    def _log_validation_result(self, result: SpreadsheetValidationResult):
        """Registra o resultado da validação no log."""
        name = result.spreadsheet_info.name
        status = result.status.value
        
        if result.status == ValidationStatus.VALID:
            self.logger.info(
                f"✓ {name}: VÁLIDA - {result.sheet_count} abas, "
                f"{result.row_count} linhas, dados: {result.has_data}"
            )
        elif result.status == ValidationStatus.WARNING:
            self.logger.warning(
                f"⚠ {name}: AVISO - {'; '.join(result.warnings)}"
            )
        elif result.status == ValidationStatus.INVALID:
            self.logger.warning(
                f"✗ {name}: INVÁLIDA - {'; '.join(result.errors)}"
            )
        else:  # ERROR
            self.logger.error(
                f"✗ {name}: ERRO - {'; '.join(result.errors)}"
            )
            
    def _log_validation_summary(self, results: List[SpreadsheetValidationResult]):
        """Registra resumo da validação no log."""
        summary = self.get_validation_summary(results)
        
        self.logger.info("=== RESUMO DA VALIDAÇÃO ===")
        self.logger.info(f"Total de arquivos: {summary['total_files']}")
        self.logger.info(f"Válidos: {summary['valid_files']}")
        self.logger.info(f"Inválidos: {summary['invalid_files']}")
        self.logger.info(f"Com erro: {summary['error_files']}")
        self.logger.info(f"Com aviso: {summary['warning_files']}")
        self.logger.info(f"Vazios: {summary['empty_files']}")
        self.logger.info(f"Taxa de sucesso: {summary['success_rate']:.1f}%")
        self.logger.info(f"Total de abas: {summary['total_sheets']}")
        self.logger.info(f"Total de linhas: {summary['total_rows']}")
        self.logger.info("=" * 30)
        
    def _check_file_size(self, file_path: str) -> tuple[list[str], list[str]]:
        """Verifica o tamanho do arquivo e retorna erros e avisos.
        
        Args:
            file_path: Caminho para o arquivo
            
        Returns:
            Tupla contendo (erros, avisos)
        """
        errors = []
        warnings = []
        
        try:
            file_size = os.path.getsize(file_path)
            
            # Verificar se o arquivo está vazio
            if file_size == 0:
                errors.append("Arquivo está vazio (0 bytes)")
                return errors, warnings
                
            # Verificar tamanho mínimo
            if file_size < 1024:  # Menor que 1KB
                errors.append("Arquivo muito pequeno para ser uma planilha válida")
                return errors, warnings
                
            # Verificar arquivo muito grande (> 50MB)
            if file_size > 50 * 1024 * 1024:  # 50MB
                warnings.append("Arquivo muito grande, pode demorar para processar")
                
        except OSError as e:
            errors.append(f"Erro ao verificar tamanho do arquivo: {str(e)}")
            
        return errors, warnings